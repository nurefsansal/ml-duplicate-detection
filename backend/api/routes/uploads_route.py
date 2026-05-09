"""
Uploads API Route - Upload listesi + dosya yükleme (normalizasyonsuz).
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import tempfile
import time
from datetime import datetime
from typing import Any

import pandas as pd
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile
from starlette.requests import Request
from pydantic import BaseModel, Field
from sqlalchemy import create_engine
from sqlalchemy import create_engine, func, text
from sqlalchemy.orm import Session, sessionmaker

from backend.models.database import ColumnMapping, NormalizationRun, NormalizedRecord, RawRecord, Upload
from backend.services.normalization_service import infer_target_field_name
from backend.api.routes.hanna_connector import ConnectorConnectionInput
from backend.services.database_connector_service import DatabaseConnectorService
from backend.services.job_service import create_job, update_job_progress
from backend.services.pipeline_log_service import (
    add_pipeline_event,
    create_pipeline_run,
    finalize_pipeline_run,
)

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql+psycopg2://ml_duplicate_user:1234@localhost:5434/ml_duplicate_db",
)
engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine)



def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _safe_val(v: Any) -> Any:
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (bool, int, float, str)):
        return v
    return str(v)


def _row_to_payload(row: dict) -> dict:
    return {str(k): _safe_val(v) for k, v in row.items()}


def _ingestion_hash(payload: dict) -> str:
    serialized = json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def _ensure_import_batch(db: Session, *, upload_id: int) -> str:
    batch_id = f"upload-{upload_id}"
    db.execute(
        text(
            """
            INSERT INTO import_batches (
                batch_id,
                source_name,
                source_type,
                status,
                record_count,
                created_at
            )
            SELECT
                :batch_id,
                uploads.file_name,
                COALESCE(uploads.source_type, 'unknown'),
                COALESCE(uploads.status, 'uploaded'),
                COALESCE(uploads.total_records, 0),
                COALESCE(uploads.created_at, CURRENT_TIMESTAMP)
            FROM uploads
            WHERE uploads.id = :upload_id
            ON CONFLICT (batch_id) DO NOTHING
            """
        ),
        {"batch_id": batch_id, "upload_id": upload_id},
    )
    return batch_id


def _suggested_target_field(source_column: str) -> str:
    target_field = infer_target_field_name(source_column)
    return "" if target_field == "ignored" else target_field


def _insert_raw_batch(
    db: Session,
    *,
    upload_id: int,
    rows: list[dict],
    start_index: int = 1,
) -> None:
    if not rows:
        return
    batch = []
    batch_id = _ensure_import_batch(db, upload_id=upload_id)
    for offset, row in enumerate(rows, start=start_index):
        payload = _row_to_payload(row)
        batch.append(
            RawRecord(
                upload_id=upload_id,
                batch_id=batch_id,
                row_index=offset,
                raw_payload=payload,
                ingestion_hash=_ingestion_hash(payload),
                row_status="pending",
            )
        )
    db.bulk_save_objects(batch)


def _iter_csv_chunks(file_path: str):
    try:
        return pd.read_csv(file_path, chunksize=2000, encoding="utf-8")
    except UnicodeDecodeError:
        return pd.read_csv(file_path, chunksize=2000, encoding="latin-1")


def _read_excel_header(file_path: str) -> list[str]:
    # Read only header row without loading the full file to memory.
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=500,
            detail=f"Excel okumak için openpyxl gerekli: {exc}",
        ) from exc

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    try:
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            return []
        header = [str(v).strip() if v is not None else "" for v in first_row]
        return [h if h else f"column_{i+1}" for i, h in enumerate(header)]
    finally:
        wb.close()


def _iter_excel_rows(file_path: str, *, chunk_size: int = 2000):
    # Stream Excel rows in chunks using openpyxl read-only mode.
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    try:
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return
        header = [str(v).strip() if v is not None else "" for v in header_row]
        header = [h if h else f"column_{i+1}" for i, h in enumerate(header)]

        batch: list[dict] = []
        for values in rows_iter:
            row_dict = {header[i]: values[i] if i < len(values) else None for i in range(len(header))}
            batch.append(row_dict)
            if len(batch) >= chunk_size:
                yield batch
                batch = []
        if batch:
            yield batch
    finally:
        wb.close()


def _run_upload_ingestion_job(
    *,
    temp_file_path: str,
    upload_id: int,
    source_type: str,
    job_id: int | None,
    request_id: str | None,
):
    """Background job: stage file -> raw_records."""
    db = SessionLocal()
    run_id: int | None = None
    t0 = time.monotonic()
    processed = 0
    try:
        run = create_pipeline_run(
            db,
            pipeline_type="upload",
            request_id=request_id,
            upload_id=upload_id,
            job_id=job_id,
            metadata={"source_type": source_type},
        )
        run_id = run.id
        add_pipeline_event(
            db,
            run_id=run.id,
            stage="ingest",
            event_type="started",
            message="Ham kayıt ingest başladı",
        )
        db.commit()

        if source_type == "csv":
            for chunk_df in _iter_csv_chunks(temp_file_path):
                records = chunk_df.to_dict(orient="records")
                processed += len(records)
                _insert_raw_batch(
                    db,
                    upload_id=upload_id,
                    rows=records,
                    start_index=processed - len(records) + 1,
                )
                db.flush()
                if job_id is not None:
                    update_job_progress(
                        db,
                        job_id=job_id,
                        status="running",
                        processed_rows=processed,
                        progress=min(99.0, float(processed % 100000) / 1000.0),
                    )
                if run_id is not None:
                    finalize_pipeline_run(db, run_id=run_id, status="running", processed_rows=processed)
                db.commit()
        else:
            for batch in _iter_excel_rows(temp_file_path, chunk_size=2000):
                processed += len(batch)
                _insert_raw_batch(
                    db,
                    upload_id=upload_id,
                    rows=batch,
                    start_index=processed - len(batch) + 1,
                )
                db.flush()
                if job_id is not None:
                    update_job_progress(
                        db,
                        job_id=job_id,
                        status="running",
                        processed_rows=processed,
                        progress=min(99.0, float(processed % 100000) / 1000.0),
                    )
                if run_id is not None:
                    finalize_pipeline_run(db, run_id=run_id, status="running", processed_rows=processed)
                db.commit()

        upload = db.query(Upload).filter(Upload.id == upload_id).first()
        if upload is not None:
            upload.total_records = int(processed)
            upload.status = "uploaded"
            upload.processing_stage = "raw"
            upload.completed_at = datetime.utcnow()
            db.flush()

        if job_id is not None:
            update_job_progress(
                db,
                job_id=job_id,
                status="completed",
                progress=100.0,
                total_rows=int(processed),
                processed_rows=int(processed),
            )

        if run_id is not None:
            duration_ms = int((time.monotonic() - t0) * 1000)
            add_pipeline_event(
                db,
                run_id=run_id,
                stage="ingest",
                event_type="completed",
                message="Ham kayıt ingest tamamlandı",
                total_rows=int(processed),
                processed_rows=int(processed),
                duration_ms=duration_ms,
            )
            finalize_pipeline_run(
                db,
                run_id=run_id,
                status="completed",
                total_rows=int(processed),
                processed_rows=int(processed),
                metadata_patch={"duration_ms": duration_ms},
            )

        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        if job_id is not None:
            update_job_progress(db, job_id=job_id, status="failed", error_message=str(exc))
        if run_id is not None:
            add_pipeline_event(db, run_id=run_id, stage="ingest", event_type="failed", message=str(exc))
            finalize_pipeline_run(db, run_id=run_id, status="failed", error_message=str(exc))
        db.commit()
    finally:
        db.close()
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except OSError:
                pass


router = APIRouter()


@router.post("/uploads/file")
async def upload_file_only(
    background_tasks: BackgroundTasks,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Dosyayı yükle: uploads oluştur, ingest arka planda raw_records'a yazar."""
    filename = (file.filename or "").lower()
    if not (
        filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")
    ):
        raise HTTPException(
            status_code=400,
            detail="Sadece .xlsx, .xls ve .csv dosyaları destekleniyor",
        )

    temp_file_path: str | None = None
    file_size_bytes = 0
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=f"_{file.filename or 'upload'}") as temp_file:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                file_size_bytes += len(chunk)
                if file_size_bytes > 100 * 1024 * 1024:
                    raise HTTPException(
                        status_code=413,
                        detail="Dosya boyutu 100 MB sınırını aşıyor. Büyük veri için CSV önerilir.",
                    )
                temp_file.write(chunk)
            temp_file_path = temp_file.name
    except HTTPException:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except OSError:
                pass
        raise
    except Exception as exc:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except OSError:
                pass
        raise HTTPException(status_code=400, detail=f"Dosya okunamadı: {exc}") from exc

    try:
        is_excel = filename.endswith(".xlsx") or filename.endswith(".xls")
        source_type = "excel" if is_excel else "csv"

        if is_excel:
            source_columns = _read_excel_header(temp_file_path)
        else:
            try:
                source_columns = list(pd.read_csv(temp_file_path, nrows=0, encoding="utf-8").columns)
            except UnicodeDecodeError:
                source_columns = list(pd.read_csv(temp_file_path, nrows=0, encoding="latin-1").columns)
            source_columns = [str(col) for col in source_columns]

        if not source_columns:
            raise HTTPException(status_code=400, detail="Dosya boş veya okunamadı")

        suggested_mappings = {col: _suggested_target_field(col) for col in source_columns}

        upload = Upload(
            source_type=source_type,
            source_name=file.filename or "uploaded_file",
            file_name=file.filename or "uploaded_file",
            file_size_bytes=file_size_bytes,
            total_records=0,
            status="processing",
            processing_stage="ingesting",
        )
        db.add(upload)
        db.flush()

        job = create_job(db, job_type="upload")
        if job is not None:
            db.flush()
        db.commit()
        db.refresh(upload)

        req_id = getattr(getattr(request, "state", None), "request_id", None)
        background_tasks.add_task(
            _run_upload_ingestion_job,
            temp_file_path=temp_file_path,
            upload_id=upload.id,
            source_type=source_type,
            job_id=job.id if job is not None else None,
            request_id=req_id,
        )
        temp_file_path = None  # background task will clean it

        return {
            "success": True,
            "upload_id": upload.id,
            "job_id": job.id if job is not None else None,
            "file_name": upload.file_name,
            "source_type": upload.source_type,
            "total_records": 0,
            "source_columns": source_columns,
            "suggested_mappings": suggested_mappings,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Upload kaydedilemedi: {exc}") from exc
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except OSError:
                pass


@router.get("/uploads/{upload_id}/columns")
def get_upload_columns(upload_id: int, db: Session = Depends(get_db)):
    """Upload'ın ham kayıt kolon listesini döndür (ilk raw_record'dan)."""
    upload = db.query(Upload).filter(Upload.id == upload_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail=f"Upload {upload_id} bulunamadı")

    first_raw = (
        db.query(RawRecord).filter(RawRecord.upload_id == upload_id).first()
    )
    if not first_raw or not first_raw.raw_payload:
        return {
            "success": True,
            "upload_id": upload_id,
            "source_columns": [],
            "suggested_mappings": {},
        }

    source_columns = list(first_raw.raw_payload.keys())
    suggested_mappings = {
        col: _suggested_target_field(col)
        for col in source_columns
    }

    return {
        "success": True,
        "upload_id": upload_id,
        "source_columns": source_columns,
        "suggested_mappings": suggested_mappings,
    }


@router.get("/uploads/{upload_id}/raw-records")
def list_raw_records(
    upload_id: int,
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db),
):
    """Ham kayıtları sayfalı listele (raw_records)."""
    upload = db.query(Upload).filter(Upload.id == upload_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail=f"Upload {upload_id} bulunamadı")

    page = max(1, int(page))
    page_size = max(1, min(500, int(page_size)))
    offset = (page - 1) * page_size

    q = db.query(RawRecord).filter(RawRecord.upload_id == upload_id)
    total = int(q.with_entities(func.count(RawRecord.id)).scalar() or 0)
    rows = (
        q.order_by(RawRecord.id.asc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    records = [
        {
            "id": r.id,
            "upload_id": r.upload_id,
            "row_index": r.row_index,
            "raw_payload": r.raw_payload or {},
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]

    return {
        "success": True,
        "upload_id": upload_id,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
        "records": records,
    }


@router.get("/uploads")
def list_uploads(
    limit: int = 50,
    has_normalized_records: bool = False,
    db: Session = Depends(get_db),
):
    """Upload listesi.

    has_normalized_records=true → sadece normalized_records tablosunda kaydı olan uploadları döner.
    Bu sayede Mükerrer Tespit dropdown'u sadece normalizasyon tamamlanmış uploadları gösterir.
    """
    try:
        q = db.query(Upload).order_by(Upload.created_at.desc())

        if has_normalized_records:
            upload_ids_with_records = [
                row[0]
                for row in db.query(NormalizedRecord.upload_id).distinct().all()
            ]
            if upload_ids_with_records:
                q = q.filter(Upload.id.in_(upload_ids_with_records))
            else:
                return {"success": True, "count": 0, "uploads": []}

        uploads = q.limit(limit).all()

        result = []
        for upload in uploads:
            latest_run = (
                db.query(NormalizationRun)
                .filter(NormalizationRun.upload_id == upload.id)
                .order_by(
                    NormalizationRun.created_at.desc(),
                    NormalizationRun.id.desc(),
                )
                .first()
            )
            result.append(
                {
                    "id": upload.id,
                    "file_name": upload.file_name,
                    "source_type": upload.source_type or "unknown",
                    "total_records": upload.total_records or 0,
                    "status": upload.status,
                    "processing_stage": upload.processing_stage,
                    "created_at": (
                        upload.created_at.isoformat() if upload.created_at else None
                    ),
                    "completed_at": (
                        upload.completed_at.isoformat()
                        if upload.completed_at
                        else None
                    ),
                    "latest_normalization_run_id": (
                        latest_run.id if latest_run else None
                    ),
                }
            )

        return {"success": True, "count": len(result), "uploads": result}
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"Error listing uploads: {exc}"
        ) from exc


class FromInstitutionUploadRequest(BaseModel):
    connection: ConnectorConnectionInput
    table: str = Field(..., min_length=1)
    limit: int | None = None


@router.post("/uploads/from-institution-db")
def upload_from_institution_db(payload: FromInstitutionUploadRequest, db: Session = Depends(get_db)):
    """Kurum veritabanından seçili tabloyu alıp uploads + raw_records oluşturur."""
    try:
        conn = payload.connection
        # Determine schema and table name
        if "." in payload.table:
            schema, table_name = payload.table.split(".", 1)
        else:
            schema = conn.db_schema or "public"
            table_name = payload.table

        # Build connector service with requested search_path
        # ensure db_schema is set so preview/selects use correct search_path
        conn.db_schema = schema
        service = DatabaseConnectorService.from_details(conn.to_details())

        # Build select SQL (apply limit if provided)
        limit_clause = f" LIMIT {int(payload.limit)}" if payload.limit else ""
        sql = f'SELECT * FROM "{table_name}"{limit_clause}'

        rows = service.fetch_rows(sql)

        if not rows:
            raise HTTPException(status_code=400, detail="Seçili tablodan veri alınamadı veya tablo boş.")

        # Create upload record
        upload = Upload(
            source_type="institution_db",
            source_name=f"{conn.label}:{schema}.{table_name}",
            file_name=f"{schema}.{table_name}",
            file_size_bytes=0,
            total_records=len(rows),
            status="uploaded",
            processing_stage="raw",
        )
        db.add(upload)
        db.flush()

        # Insert raw records in chunks
        BATCH = 2000
        batch = []
        for r in rows:
            batch.append(r)
            if len(batch) >= BATCH:
                _insert_raw_batch(db, upload_id=upload.id, rows=batch)
                db.flush()
                batch = []
        if batch:
            _insert_raw_batch(db, upload_id=upload.id, rows=batch)

        db.commit()
        db.refresh(upload)
        return {
            "success": True,
            "upload_id": upload.id,
            "source": upload.source_name,
            "total_records": upload.total_records,
        }

    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Institution import failed: {exc}") from exc
